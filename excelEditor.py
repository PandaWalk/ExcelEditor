import openpyxl as x

wb = x.load_workbook('/Users/mac_12/Desktop/Aktionspakete_2025-3_20250127155206.xltx')


def number(sh):
    sh = wb.active
    sh.insert_cols(1)

    for n in range(1, sh.max_row):
        celval = n
        cell = sh.cell(n, 1)
        cell.value = celval

#zählt die Reihen
def countrow(sh):
    sh = wb.active
    counter = 0
    #print(counter)
    for n in range(1, sh.max_row + 1):
        #print(counter)
        if sh.cell(2, n).value != 1 and sh.cell(2, n).value is not None:
            break

        counter += 1

    return counter

# Nrhs80#6

def countDupes(sh):
    sh = wb.active
    maxcol = countrow(sh)
    dupecounter = 0
    worklist = []
    dupelist = []

    for n in range(2, sh.max_row + 1):

        for m in range(1, maxcol + 6):

            if sh.cell(n, m).value == 1:
                dupecounter += 1
                worklist.append([n, m])

            if sh.cell(n, m) == sh.cell(n, maxcol) and dupecounter <= 1:
                dupecounter = 0
                worklist.clear()

            if sh.cell(n, m) == sh.cell(n, maxcol) and dupecounter > 1:
                dupecounter = 0
                c = 0
                for element in worklist:
                    dupelist.append(worklist[c])
                    c += 1
                worklist.clear()

    return dupelist


# Nimmt 2 Zellen als parameter entgegen
def splitDupes():
    sh = wb.active
    alldupes = []
    alldupes = countDupes(sh)

    worklist = []

    for n in range(2, sh.max_row + 1):
        worklist.clear()

        for element in alldupes:

            if element[0] == n:
                worklist.append(element)

            if element[0] != n and element == alldupes[-1] or element[0] == n and element == alldupes[-1]:

                valuelist = []
                for elements in worklist:
                    splitted = str(sh.cell(1, elements[1]).value).split(':')
                    valuelist.append(splitted[1])

    return worklist


def sort():
    sh = wb.active
    maxcol = countrow(sh)

    # Finde die Summenzeile und die Zeile darunter
    summen_zeile = None
    summen_daten = None
    nach_summen_zeile = None
    nach_summen_daten = None

    # Listen für Zeilen mit Zahlen größer als 1
    grosse_zahlen_zeilen = []
    grosse_zahlen_daten = []

    for row in range(2, sh.max_row + 1):
        # Überprüfe die erste Spalte auf "Summe"
        if sh.cell(row, 1).value and "Summe" in str(sh.cell(row, 1).value):
            summen_zeile = row
            # Speichere die Summendaten
            summen_daten = [sh.cell(summen_zeile, col).value for col in range(1, sh.max_column + 1)]
            # Speichere auch die Daten der Zeile nach der Summe
            if row + 1 <= sh.max_row:
                nach_summen_zeile = row + 1
                nach_summen_daten = [sh.cell(nach_summen_zeile, col).value for col in range(1, sh.max_column + 1)]
            break

    # Finde Zeilen mit Zahlen größer als 1
    for row in range(2, sh.max_row + 1):
        if row != summen_zeile and row != nach_summen_zeile:
            hat_grosse_zahl = False
            for col in range(1, maxcol + 1):
                cell_value = sh.cell(row, col).value
                if isinstance(cell_value, (int, float)) and cell_value > 1:
                    hat_grosse_zahl = True
                    break
            if hat_grosse_zahl:
                grosse_zahlen_zeilen.append(row)
                grosse_zahlen_daten.append([sh.cell(row, col).value for col in range(1, sh.max_column + 1)])

    # Startzeile für die Sortierung (erste Datenzeile)
    start_row = 2

    # Durchlaufe jede Spalte bis zur maximalen Spalte
    for col in range(1, maxcol + 1):
        column_values = []

        # Sammle alle Werte der aktuellen Spalte (nur Einsen)
        for row in range(start_row, sh.max_row + 1):
            if (row != summen_zeile and
                    row != nach_summen_zeile and
                    row not in grosse_zahlen_zeilen):
                cell_value = sh.cell(row, col).value
                if cell_value == 1:
                    column_values.append(row)

        # Sortiere die Werte treppenförmig
        current_row = start_row
        for row_index in column_values:
            if row_index != current_row:
                # Tausche die Zeilen
                for c in range(1, sh.max_column + 1):
                    temp_value = sh.cell(current_row, c).value
                    sh.cell(current_row, c).value = sh.cell(row_index, c).value
                    sh.cell(row_index, c).value = temp_value
            current_row += 1

        # Aktualisiere die Startzeile für die nächste Spalte
        start_row = current_row

    # Berechne die Position, wo die Zeilen mit großen Zahlen eingefügt werden sollen
    insert_row = start_row

    # Füge die Zeilen mit großen Zahlen ein
    for i, zeilen_daten in enumerate(grosse_zahlen_daten):
        for col in range(1, sh.max_column + 1):
            sh.cell(insert_row + i, col).value = zeilen_daten[col - 1]

    # Füge die Summenzeile als vorletzte Zeile ein
    if summen_daten:
        for col in range(1, sh.max_column + 1):
            sh.cell(sh.max_row - 1, col).value = summen_daten[col - 1]

    # Füge die Zeile nach der Summe als letzte Zeile ein
    if nach_summen_daten:
        for col in range(1, sh.max_column + 1):
            sh.cell(sh.max_row, col).value = nach_summen_daten[col - 1]

def createHeaderlist():
    sh = wb.active
    headerlist = []
    # Durchsuche die erste Zeile nach Headers, die "Paket" enthalten
    for col in range(1, sh.max_column + 1):
        cell_value = sh.cell(1, col).value
        if cell_value and "Paket" in str(cell_value):
            headerlist.append(str(cell_value))
    return headerlist


def createNewHeaders(sh):
    sh = wb.active
    maxcol = countrow(sh)
    dupes = countDupes(sh)
    existing_headers = createHeaderlist()
    new_headers = []
    header_mapping = {}

    # Finde die Summenzeile
    sum_row = 2  # Standardwert, falls keine Summenzeile gefunden wird
    for row in range(2, sh.max_row + 1):
        if str(sh.cell(row, 1).value).startswith('Summe'):
            sum_row = row
            break

    # Finde die höchste existierende Paketnummer
    highest_package_number = 0
    for header in existing_headers:
        try:
            number = int(header.split(':')[0].replace('Paket', '').strip())
            highest_package_number = max(highest_package_number, number)
        except ValueError:
            continue

    # Prüfe, welche Spaltenkombinationen tatsächlich mit 1en existieren
    valid_column_combinations = set()
    for row in range(2, sum_row - 1):  # Nur bis VOR der Summenzeile
        current_cols = []
        for col in range(1, maxcol + 1):
            if sh.cell(row, col).value == 1:
                current_cols.append(col)
        if len(current_cols) > 1:
            valid_column_combinations.add(tuple(sorted(current_cols)))

    # Zähle wie viele neue Header wir erstellen werden
    new_header_count = len(valid_column_combinations)

    # Füge alle neuen Spalten auf einmal ein
    if new_header_count > 0:
        sh.insert_cols(maxcol + 1, new_header_count)

    # Erstelle neue Header und verarbeite die Zeilen
    current_new_col = maxcol + 1
    new_package_number = highest_package_number + 1

    for cols in valid_column_combinations:
        header_names = []
        for col in cols:
            header_value = sh.cell(1, col).value
            if header_value:
                header_name = str(header_value).split(':')[-1].strip()
                header_names.append(header_name)

        combined_names = " + ".join(header_names)
        new_header = f"Paket{new_package_number}:{combined_names}"

        if new_header not in existing_headers and new_header not in new_headers:
            new_headers.append(new_header)
            sh.cell(1, current_new_col).value = new_header

            # Verarbeite die Zeilen: Lösche die alten 1en und setze die neue 1
            for r in range(2, sum_row - 1):  # Nur bis VOR der Summenzeile
                if all(sh.cell(r, c).value == 1 for c in cols):
                    # Lösche die alten 1en
                    for c in cols:
                        sh.cell(r, c).value = None
                    # Setze die neue 1 in der neuen Spalte
                    sh.cell(r, current_new_col).value = 1

            # Aktualisiere die Summenzeile
            sum_formula = f"=SUM({sh.cell(2, current_new_col).coordinate}:{sh.cell(sum_row - 2, current_new_col).coordinate})"
            sh.cell(sum_row - 1, current_new_col).value = sum_formula

            current_new_col += 1
            new_package_number += 1

    return new_headers


def main():
    new_headers = createNewHeaders(wb.active)
    sort()
    print("Neue Header erstellt:", new_headers)
    wb.save('/Users/mac_12/Desktop/test.xls')



if __name__ == '__main__':
    main()
